#!/usr/bin/env python3
"""Build the cumulative BIOMAP workbook from native BioAnalyst runs.

The cumulative workbook is regenerated on each run by scanning every valid
forecast directory under `model_forecast`. This keeps one historical Excel
artifact updated over time while the per-run cell-matrix workbooks remain
separate outputs.
"""

from __future__ import annotations

import argparse
import json
import math
from collections import OrderedDict
from pathlib import Path
from typing import Any, Callable

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd

from biomap_curated_features import CURATED_BIOMAP_FEATURES

REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_MODEL_FORECAST_ROOT = REPO_ROOT / "outputs" / "local_preview" / "model_forecast"
DEFAULT_PREVISIONI_DIRNAME = "previsioni"
DEFAULT_CHART_DIRNAME = "chart"
DEFAULT_AREA_ANALYSIS_DIRNAME = "area_analysis"
DEFAULT_WORKBOOK_NAME = "FINAL_FEATURES_ANALYSIS.xlsx"

import analyze_latest_native_tests as ant  # noqa: E402
from biomap_metric_utils import PRECIPITATION_AUDIT, continuous_metric_summary  # noqa: E402


READY_STATES = {
    "ready_high",
    "usable_medium",
    "diagnostic_medium",
    "diagnostic_promising",
    "context_spatially_consistent",
    "diagnostic_spatially_consistent",
}


VARIABLE_SHEET_COLUMNS = [
    "run_dir",
    "forecast_month",
    "checkpoint",
    "input_mode",
    "group",
    "variable",
    "level_index",
    "pressure_level",
    "unit",
    "source_status",
    "reliability",
    "cell_count",
    "valid_cell_count",
    "predicted_mean",
    "observed_mean",
    "mae",
    "rmse",
    "bias",
    "correlation",
    "predicted_min",
    "predicted_max",
    "observed_min",
    "observed_max",
    "wape_pct",
    "smaape_pct",
    "smape_pct",
    "rmae_pct",
    "cvrmse_pct",
    "source_report_year",
    "source_report_path",
    "forecast_month_dt",
    "year",
    "month",
    "month_name",
    "feature_key",
    "r2_corr_squared",
    "biomap_domain",
    "biomap_use",
    "biomap_interpretation",
    "biomap_readiness",
]


FEATURE_METADATA = {
    ("climate", "t2m"): {
        "biomap_domain": "temperature_stress",
        "biomap_use": "core_forecast_indicator",
        "biomap_interpretation": "Damage/restoration climate pressure",
        "thesis_role": "primary forecast indicator",
        "recommended_use_in_thesis": "Usare come indicatore previsionale principale di stress climatico.",
        "main_caution": "Attenzione al bias stagionale, soprattutto nei mesi freddi.",
        "next_step": "Validazione estesa fuori training e test 2020-2026.",
    },
    ("climate", "tp"): {
        "biomap_domain": "precipitation_water_stress",
        "biomap_use": "secondary_diagnostic",
        "biomap_interpretation": "Hydrological stress and drought/wetness context",
        "thesis_role": "supporting forecast indicator",
        "recommended_use_in_thesis": "Usare come indicatore idrico di supporto, non come variabile unica.",
        "main_caution": "La precipitazione e piu instabile e sensibile agli estremi.",
        "next_step": "Aggiungere test evento-specifici e generalizzazione fuori periodo.",
    },
    ("vegetation", "NDVI"): {
        "biomap_domain": "vegetation_condition",
        "biomap_use": "promising_diagnostic",
        "biomap_interpretation": "Restoration greenness / vegetation condition",
        "thesis_role": "ecological support indicator",
        "recommended_use_in_thesis": "Usare come supporto ecologico per stato della vegetazione.",
        "main_caution": "Richiede coerenza temporale del dataset NDVI esteso.",
        "next_step": "Validare mesi e anni aggiuntivi con dati osservati estesi.",
    },
    ("forest", "Forest"): {
        "biomap_domain": "forest_context",
        "biomap_use": "context_layer",
        "biomap_interpretation": "Forest damage/restoration context",
        "thesis_role": "contextual indicator",
        "recommended_use_in_thesis": "Usare come layer contestuale per restoration/damage.",
        "main_caution": "Layer poco dinamico nel tempo; leggerlo come contesto.",
        "next_step": "Estendere la copertura annuale e validare la coerenza spaziale.",
    },
    ("agriculture", "Agriculture"): {
        "biomap_domain": "agriculture_pressure",
        "biomap_use": "context_layer",
        "biomap_interpretation": "Anthropic land pressure",
        "thesis_role": "contextual pressure indicator",
        "recommended_use_in_thesis": "Usare come pressione antropica di contesto.",
        "main_caution": "Variabile spesso piu statica che previsionale.",
        "next_step": "Chiarire la frequenza temporale del dataset sorgente.",
    },
    ("agriculture", "Cropland"): {
        "biomap_domain": "cropland_pressure",
        "biomap_use": "context_layer",
        "biomap_interpretation": "Anthropic land pressure",
        "thesis_role": "contextual pressure indicator",
        "recommended_use_in_thesis": "Usare per stimare pressione agricola e frammentazione.",
        "main_caution": "Interpretare come layer di pressione, non come target climatico.",
        "next_step": "Verificare consistenza annuale e copertura 2020-2026.",
    },
    ("agriculture", "Arable"): {
        "biomap_domain": "arable_pressure",
        "biomap_use": "context_layer",
        "biomap_interpretation": "Anthropic land pressure",
        "thesis_role": "contextual pressure indicator",
        "recommended_use_in_thesis": "Usare come contesto agricolo complementare.",
        "main_caution": "Puo essere quasi statico nel tempo.",
        "next_step": "Documentare la frequenza di aggiornamento del layer.",
    },
    ("atmospheric", "q"): {
        "biomap_domain": "atmospheric_moisture",
        "biomap_use": "diagnostic",
        "biomap_interpretation": "Atmospheric moisture support",
        "thesis_role": "diagnostic climate variable",
        "recommended_use_in_thesis": "Usare come diagnostica atmosferica di supporto.",
        "main_caution": "La lettura dipende dal livello di pressione usato nel curated export.",
        "next_step": "Documentare esplicitamente il level_index nel workbook finale.",
    },
    ("edaphic", "swvl1"): {
        "biomap_domain": "soil_water_stress",
        "biomap_use": "diagnostic",
        "biomap_interpretation": "Shallow soil moisture support",
        "thesis_role": "ecohydrological support indicator",
        "recommended_use_in_thesis": "Usare come indicatore di stress idrico superficiale.",
        "main_caution": "Serve confronto stagionale coerente per letture corrette.",
        "next_step": "Estendere audit dei dati e validazione 2020-2026.",
    },
    ("edaphic", "swvl2"): {
        "biomap_domain": "soil_water_stress",
        "biomap_use": "diagnostic",
        "biomap_interpretation": "Deeper soil moisture support",
        "thesis_role": "ecohydrological support indicator",
        "recommended_use_in_thesis": "Usare come indicatore di stress idrico piu profondo.",
        "main_caution": "Puo reagire in modo piu lento rispetto a swvl1.",
        "next_step": "Confrontare prestazioni per stagioni e aree climatiche diverse.",
    },
    ("edaphic", "stl1"): {
        "biomap_domain": "soil_temperature",
        "biomap_use": "diagnostic",
        "biomap_interpretation": "Shallow soil temperature support",
        "thesis_role": "soil temperature indicator",
        "recommended_use_in_thesis": "Usare per leggere la temperatura del suolo superficiale.",
        "main_caution": "Richiede confronto stagionale coerente e contesto climatico.",
        "next_step": "Confrontare il segnale con t2m e con il contesto del suolo.",
    },
    ("edaphic", "stl2"): {
        "biomap_domain": "soil_temperature",
        "biomap_use": "diagnostic",
        "biomap_interpretation": "Deeper soil temperature support",
        "thesis_role": "soil temperature indicator",
        "recommended_use_in_thesis": "Usare per leggere la temperatura del suolo piu profonda.",
        "main_caution": "Puo rispondere con inerzia maggiore rispetto a stl1.",
        "next_step": "Analizzare differenze temporali tra stl1 e stl2.",
    },
    ("climate", "avg_snlwrf"): {
        "biomap_domain": "surface_radiation_balance",
        "biomap_use": "diagnostic",
        "biomap_interpretation": "Surface net long-wave radiation support",
        "thesis_role": "radiation diagnostic variable",
        "recommended_use_in_thesis": "Usare come variabile radiativa di supporto ai trend climatici.",
        "main_caution": "Piu difficile da comunicare da sola come indicatore finale.",
        "next_step": "Legarla ai trend di temperatura, umidita e suolo nelle analisi.",
    },
    ("land", "Land"): {
        "biomap_domain": "land_mask_context",
        "biomap_use": "technical_context",
        "biomap_interpretation": "Mask/support variable",
        "thesis_role": "technical support variable",
        "recommended_use_in_thesis": "Usare come supporto tecnico o maschera.",
        "main_caution": "Non interpretare come indicatore ecologico diretto.",
        "next_step": "Lasciare come layer tecnico.",
    },
    ("redlist", "RLI"): {
        "biomap_domain": "redlist_context",
        "biomap_use": "context_layer",
        "biomap_interpretation": "Conservation-risk context",
        "thesis_role": "conservation context indicator",
        "recommended_use_in_thesis": "Usare come contesto di rischio conservazionistico.",
        "main_caution": "Indice non mensile e poco adatto a forecasting dinamico.",
        "next_step": "Gestirlo come layer di contesto non mensile.",
    },
    ("misc", "avg_slhtf"): {
        "biomap_domain": "energy_balance_context",
        "biomap_use": "diagnostic",
        "biomap_interpretation": "Surface latent heat flux support",
        "thesis_role": "diagnostic support variable",
        "recommended_use_in_thesis": "Usare solo come diagnostica climatica di supporto.",
        "main_caution": "Difficile da comunicare come indicatore finale.",
        "next_step": "Valutare se mantenerlo solo nel workbook tecnico.",
    },
    ("misc", "avg_pevr"): {
        "biomap_domain": "evaporation_context",
        "biomap_use": "diagnostic",
        "biomap_interpretation": "Potential evaporation rate support",
        "thesis_role": "diagnostic support variable",
        "recommended_use_in_thesis": "Usare come supporto ai ragionamenti su bilancio idrico.",
        "main_caution": "Non usarlo da solo come evidenza finale.",
        "next_step": "Legarlo a tp e soil moisture nei test futuri.",
    },
}


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Aggiorna il workbook finale BIOMAP dal run BioAnalyst native.")
    parser.add_argument("--run-dir", type=Path, default=None)
    parser.add_argument("--runs-root", type=Path, default=DEFAULT_MODEL_FORECAST_ROOT)
    parser.add_argument("--workbook-path", type=Path, default=None)
    return parser


def load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def safe_timestamp(value: str | None) -> pd.Timestamp | None:
    if not value:
        return None
    try:
        return pd.Timestamp(value)
    except Exception:
        return None


def run_label(manifest: dict[str, Any], run_dir: Path) -> str:
    return str(manifest.get("label") or run_dir.name)


def infer_runs_root(
    current_run_dir: Path | None,
    workbook_path: Path | None,
    fallback_runs_root: Path = DEFAULT_MODEL_FORECAST_ROOT,
) -> Path:
    if current_run_dir is not None:
        for candidate in [current_run_dir, *current_run_dir.parents]:
            if candidate.name.lower() == "model_forecast":
                return candidate
    if workbook_path is not None:
        parent = workbook_path.parent
        if parent.name == DEFAULT_PREVISIONI_DIRNAME:
            return parent.parent
        if (parent / "model_forecast").exists():
            return parent / "model_forecast"
    return fallback_runs_root


def output_root_from_runs_root(runs_root: Path) -> Path:
    if runs_root.name.lower() == "model_forecast":
        return runs_root.parent
    return runs_root


def default_workbook_path(runs_root: Path) -> Path:
    return output_root_from_runs_root(runs_root) / DEFAULT_WORKBOOK_NAME


def default_chart_root(runs_root: Path) -> Path:
    return output_root_from_runs_root(runs_root) / DEFAULT_CHART_DIRNAME


def default_area_analysis_root(runs_root: Path) -> Path:
    return output_root_from_runs_root(runs_root) / DEFAULT_AREA_ANALYSIS_DIRNAME


def iter_run_dirs(root: Path) -> list[Path]:
    run_dirs: list[Path] = []
    for manifest_path in sorted(root.rglob("forecast_native_manifest.json")):
        run_dir = manifest_path.parent
        parts = {part.lower() for part in run_dir.parts}
        if DEFAULT_PREVISIONI_DIRNAME.lower() in parts:
            continue
        if any("sintesi_affidabilita" in part.lower() for part in run_dir.parts):
            continue
        if any("analysis_report" in part.lower() for part in run_dir.parts):
            continue
        if not (run_dir / "native_prediction_original.pt").exists():
            continue
        if not (run_dir / "native_target_original.pt").exists():
            continue
        run_dirs.append(run_dir)
    unique = list(dict.fromkeys(run_dirs))
    unique.sort(
        key=lambda path: (
            safe_timestamp(load_json(path / "forecast_native_manifest.json").get("forecast_month")) or pd.Timestamp.max,
            path.name.lower(),
        )
    )
    return unique


def metric_metadata(group: str, variable: str) -> dict[str, str]:
    default = {
        "biomap_domain": "native_bioanalyst_variable",
        "biomap_use": "exploratory_native",
        "biomap_interpretation": "Use only after variable-specific validation",
        "thesis_role": "exploratory variable",
        "recommended_use_in_thesis": "Usare con cautela solo dopo validazione specifica.",
        "main_caution": "Variabile non ancora prioritaria o non ancora ben validata.",
        "next_step": "Raccogliere piu test e chiarire copertura dei dati sorgente.",
    }
    if group == "species":
        default.update(
            {
                "biomap_domain": "species_distribution_proxy",
                "biomap_use": "exploratory_species",
                "biomap_interpretation": "Native species distribution channel",
                "thesis_role": "experimental biodiversity variable",
                "recommended_use_in_thesis": "Usare come layer sperimentale, non come indicatore finale unico.",
                "main_caution": "Le specie sono la parte meno stabile del modello.",
                "next_step": "Aggiungere lookup tassonomica e ulteriori metriche ecologiche.",
            }
        )
    if group == "species_proxy":
        default.update(
            {
                "biomap_domain": "biodiversity_proxy",
                "biomap_use": "exploratory",
                "biomap_interpretation": "Native species proxy, not final richness",
                "thesis_role": "biodiversity proxy",
                "recommended_use_in_thesis": "Usare come proxy esplorativo di biodiversita.",
                "main_caution": "Non presentare come richness finale validata.",
                "next_step": "Validazione piu rigorosa prima di ogni uso conclusivo.",
            }
        )
    return FEATURE_METADATA.get((group, variable), default)


def readiness(group: str, variable: str, mae: float, bias: float, corr: float, rel_mae: float) -> str:
    if (group, variable) == ("climate", "t2m"):
        if mae <= 2.0 and abs(bias) <= 1.0 and corr >= 0.90:
            return "ready_high"
        if mae <= 3.5 and abs(bias) <= 3.0 and corr >= 0.75:
            return "usable_medium"
        return "needs_calibration"
    if (group, variable) == ("climate", "tp"):
        if corr >= 0.60 and rel_mae <= 100 and abs(bias) <= 2.0:
            return "diagnostic_medium"
        return "weak_precipitation"
    if group == "species_proxy":
        return "exploratory_not_final"
    if group == "species":
        return "exploratory_species_map"
    if (group, variable) == ("vegetation", "NDVI"):
        return "diagnostic_promising" if corr >= 0.70 else "diagnostic_uncertain"
    if group in {"forest", "agriculture", "redlist"}:
        return "context_spatially_consistent" if corr >= 0.70 else "context_uncertain"
    if group in {"edaphic", "misc", "atmospheric"}:
        return "diagnostic_spatially_consistent" if corr >= 0.75 else "diagnostic_uncertain"
    if group == "climate":
        return "diagnostic_spatially_consistent" if corr >= 0.75 else "diagnostic_uncertain"
    return "native_metric_only"


def add_biomap_columns(frame: pd.DataFrame) -> pd.DataFrame:
    output = frame.copy()
    domains = []
    uses = []
    interpretations = []
    readiness_values = []
    for _, row in output.iterrows():
        meta = metric_metadata(str(row["group"]), str(row["variable"]))
        domains.append(meta["biomap_domain"])
        uses.append(meta["biomap_use"])
        interpretations.append(meta["biomap_interpretation"])
        readiness_values.append(
            readiness(
                str(row["group"]),
                str(row["variable"]),
                float(row.get("mae", math.nan)),
                float(row.get("bias", math.nan)),
                float(row.get("correlation", math.nan)),
                float(row.get("rmae_pct", row.get("relative_mae_pct", math.nan))),
            )
        )
    output["biomap_domain"] = domains
    output["biomap_use"] = uses
    output["biomap_interpretation"] = interpretations
    output["biomap_readiness"] = readiness_values
    return output


def binary_scores(pred: np.ndarray, obs: np.ndarray) -> dict[str, float | int]:
    tp = int(np.logical_and(pred, obs).sum())
    fp = int(np.logical_and(pred, ~obs).sum())
    fn = int(np.logical_and(~pred, obs).sum())
    tn = int(np.logical_and(~pred, ~obs).sum())
    precision = math.nan if tp + fp == 0 else tp / (tp + fp)
    recall = math.nan if tp + fn == 0 else tp / (tp + fn)
    f1 = math.nan if 2 * tp + fp + fn == 0 else (2 * tp) / (2 * tp + fp + fn)
    jaccard = math.nan if tp + fp + fn == 0 else tp / (tp + fp + fn)
    specificity = math.nan if tn + fp == 0 else tn / (tn + fp)
    accuracy = math.nan if tp + tn + fp + fn == 0 else (tp + tn) / (tp + tn + fp + fn)
    return {
        "tp": tp,
        "fp": fp,
        "fn": fn,
        "tn": tn,
        "precision": precision,
        "recall": recall,
        "f1_score": f1,
        "jaccard_similarity": jaccard,
        "sorensen_similarity": f1,
        "accuracy": accuracy,
        "specificity": specificity,
    }


def continuous_scores(
    pred: np.ndarray,
    obs: np.ndarray,
    *,
    group: str | None = None,
    variable: str | None = None,
) -> dict[str, float | int]:
    return continuous_metric_summary(pred, obs, group=group, variable=variable)


def load_all_metrics(runs: list[Path], current_batches: dict[Path, tuple[Any, Any, dict[str, Any]]] | None = None) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    cached = current_batches or {}
    for run_dir in runs:
        if run_dir in cached:
            prediction_batch, observed_batch, manifest = cached[run_dir]
        else:
            manifest = load_json(run_dir / "forecast_native_manifest.json")
            prediction_batch = ant.load_batch(run_dir / "native_prediction_original.pt")
            observed_batch = ant.load_batch(run_dir / "native_target_original.pt")
        rows.extend(ant.variable_metrics_for_run(run_dir, prediction_batch, observed_batch, manifest))
    frame = pd.DataFrame(rows)
    if frame.empty:
        return pd.DataFrame(columns=VARIABLE_SHEET_COLUMNS)
    frame = add_biomap_columns(frame)
    frame["forecast_month_dt"] = pd.to_datetime(frame["forecast_month"], errors="coerce")
    frame["year"] = frame["forecast_month_dt"].dt.year
    frame["month"] = frame["forecast_month_dt"].dt.month
    frame["month_name"] = frame["forecast_month_dt"].dt.strftime("%b")
    frame["feature_key"] = frame["group"].astype(str) + "." + frame["variable"].astype(str)
    frame["r2_corr_squared"] = frame["correlation"] ** 2
    frame["source_report_year"] = frame["year"]
    frame["source_report_path"] = frame["run_dir"].map(lambda item: str(Path(item) / "forecast_native_manifest.json"))
    for column in VARIABLE_SHEET_COLUMNS:
        if column not in frame.columns:
            frame[column] = None
    frame = frame[VARIABLE_SHEET_COLUMNS].sort_values(
        ["forecast_month_dt", "group", "variable", "level_index"],
        na_position="last",
    )
    return frame.reset_index(drop=True)


def compute_species_summary(runs: list[Path], current_batches: dict[Path, tuple[Any, Any, dict[str, Any]]] | None = None) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    cached = current_batches or {}
    for run_dir in runs:
        if run_dir in cached:
            pred_batch, obs_batch, manifest = cached[run_dir]
        else:
            manifest = load_json(run_dir / "forecast_native_manifest.json")
            pred_batch = ant.load_batch(run_dir / "native_prediction_original.pt")
            obs_batch = ant.load_batch(run_dir / "native_target_original.pt")
        lat_flip, lon_flip = ant.prediction_alignment_flags(manifest)
        pred_stack = ant.species_stack(pred_batch)
        obs_stack = ant.species_stack(obs_batch)
        pred_stack = np.stack(
            [ant.align_prediction(item, latitude_flip=lat_flip, longitude_flip=lon_flip) for item in pred_stack],
            axis=0,
        )
        forecast_ts = pd.Timestamp(manifest.get("forecast_month"))
        for threshold in [0.1, 0.5]:
            pred_bin = pred_stack >= threshold
            obs_bin = obs_stack >= threshold
            pred_richness = pred_bin.sum(axis=0)
            obs_richness = obs_bin.sum(axis=0)
            binary = binary_scores(pred_bin, obs_bin)
            richness = continuous_scores(
                pred_richness.reshape(-1),
                obs_richness.reshape(-1),
                group="species_proxy",
                variable=f"richness_threshold_{threshold}",
            )
            rows.append(
                {
                    "forecast_month": manifest.get("forecast_month"),
                    "year": int(forecast_ts.year),
                    "month": int(forecast_ts.month),
                    "month_name": forecast_ts.strftime("%b"),
                    "threshold": threshold,
                    "tp": binary["tp"],
                    "fp": binary["fp"],
                    "fn": binary["fn"],
                    "tn": binary["tn"],
                    "precision": binary["precision"],
                    "recall": binary["recall"],
                    "f1_score": binary["f1_score"],
                    "jaccard_similarity": binary["jaccard_similarity"],
                    "sorensen_similarity": binary["sorensen_similarity"],
                    "accuracy": binary["accuracy"],
                    "specificity": binary["specificity"],
                    "richness_predicted_mean": richness["predicted_mean"],
                    "richness_observed_mean": richness["observed_mean"],
                    "richness_mae": richness["mae"],
                    "richness_rmse": richness["rmse"],
                    "richness_bias": richness["bias"],
                    "richness_correlation": richness["correlation"],
                    "richness_wape_pct": richness["wape_pct"],
                    "richness_smaape_pct": richness["smaape_pct"],
                    "richness_smape_pct": richness["smape_pct"],
                    "richness_rmae_pct": richness["rmae_pct"],
                    "richness_cvrmse_pct": richness["cvrmse_pct"],
                    "richness_cell_count": richness["cell_count"],
                    "biomap_readiness": "exploratory_not_final",
                    "biomap_use": "exploratory_species",
                    "run_dir": str(run_dir),
                }
            )
    frame = pd.DataFrame(rows)
    if not frame.empty:
        frame = frame.sort_values(["forecast_month", "threshold"], na_position="last").reset_index(drop=True)
    return frame


def build_species_index(metrics: pd.DataFrame) -> pd.DataFrame:
    species_frame = metrics.loc[metrics["group"] == "species", ["variable"]].dropna().drop_duplicates()
    if species_frame.empty:
        return pd.DataFrame(columns=["species_id", "sheet_name", "scientific_name", "common_name", "note"])
    species_frame = species_frame.sort_values("variable").reset_index(drop=True)
    rows = [
        {
            "species_id": str(row["variable"]),
            "sheet_name": f"sp_{row['variable']}",
            "scientific_name": None,
            "common_name": None,
            "note": "ID specie del canale target BioAnalyst",
        }
        for _, row in species_frame.iterrows()
    ]
    return pd.DataFrame(rows)


def build_dashboard(metrics: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for keys, group_frame in metrics.groupby(["group", "variable", "unit", "biomap_domain", "biomap_use"], dropna=False):
        ready_months = group_frame.loc[group_frame["biomap_readiness"].isin(READY_STATES), "forecast_month"].dropna().unique()
        n_months = int(group_frame["forecast_month"].dropna().nunique())
        rows.append(
            {
                "group": keys[0],
                "variable": keys[1],
                "unit": keys[2],
                "biomap_domain": keys[3],
                "biomap_use": keys[4],
                "n_months": n_months,
                "predicted_mean": float(group_frame["predicted_mean"].mean()),
                "observed_mean": float(group_frame["observed_mean"].mean()),
                "mean_mae": float(group_frame["mae"].mean()),
                "mean_rmse": float(group_frame["rmse"].mean()),
                "mean_bias": float(group_frame["bias"].mean()),
                "mean_abs_bias": float(np.nanmean(np.abs(group_frame["bias"]))),
                "mean_corr": float(group_frame["correlation"].mean()),
                "median_corr": float(group_frame["correlation"].median()),
                "mean_r2_corr_squared": float(group_frame["r2_corr_squared"].mean()),
                "mean_wape_pct": float(group_frame["wape_pct"].mean()),
                "mean_smape_pct": float(group_frame["smape_pct"].mean()),
                "mean_rmae_pct": float(group_frame["rmae_pct"].mean()),
                "mean_cvrmse_pct": float(group_frame["cvrmse_pct"].mean()),
                "ready_or_usable_months": int(len(ready_months)),
                "readiness_ratio": math.nan if n_months == 0 else float(len(ready_months) / n_months),
            }
        )
    frame = pd.DataFrame(rows)
    if not frame.empty:
        frame = frame.sort_values(["readiness_ratio", "mean_corr"], ascending=[False, False], na_position="last").reset_index(drop=True)
    return frame


def validation_state(row: pd.Series) -> str:
    if row["readiness_ratio"] >= 0.8:
        return "high_or_context_ready"
    if row["readiness_ratio"] >= 0.4:
        return "medium_promising"
    if row["biomap_use"] in {"context_layer", "diagnostic", "technical_context"}:
        return "context_or_diagnostic"
    return "exploratory"


def build_indicator_map(dashboard: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for _, row in dashboard.iterrows():
        meta = metric_metadata(str(row["group"]), str(row["variable"]))
        rows.append(
            {
                "group": row["group"],
                "variable": row["variable"],
                "biomap_domain": row["biomap_domain"],
                "biomap_use": row["biomap_use"],
                "thesis_role": meta["thesis_role"],
                "validation_state": validation_state(row),
                "n_months": row["n_months"],
                "mean_mae": row["mean_mae"],
                "mean_rmse": row["mean_rmse"],
                "mean_wape_pct": row["mean_wape_pct"],
                "mean_smape_pct": row["mean_smape_pct"],
                "mean_rmae_pct": row["mean_rmae_pct"],
                "mean_cvrmse_pct": row["mean_cvrmse_pct"],
                "mean_bias": row["mean_bias"],
                "mean_corr": row["mean_corr"],
                "readiness_ratio": row["readiness_ratio"],
                "recommended_use_in_thesis": meta["recommended_use_in_thesis"],
                "main_caution": meta["main_caution"],
                "next_step": meta["next_step"],
            }
        )
    frame = pd.DataFrame(rows)
    if not frame.empty:
        frame = frame.sort_values(["readiness_ratio", "mean_corr"], ascending=[False, False], na_position="last").reset_index(drop=True)
    return frame


def metric_guide() -> pd.DataFrame:
    return pd.DataFrame(
        [
            ("MAE", "Media dell'errore assoluto tra predetto e osservato.", "Variabili continue", "Non distingue la direzione dell'errore."),
            ("RMSE", "Radice della media degli errori quadratici.", "Variabili continue", "Pesa di piu gli errori grandi."),
            ("bias", "Media di (predetto - osservato).", "Variabili continue", "Positivo = sovrastima; negativo = sottostima."),
            ("correlation", "Correlazione di Pearson tra predetto e osservato.", "Pattern spaziali o strutturali", "Alta correlazione non implica valori assoluti corretti."),
            ("WAPE", "100 * somma(|predetto - osservato|) / somma(|osservato|).", "Errore percentuale aggregato", "Non definito se la somma degli osservati assoluti e zero."),
            ("sMAPE", "Media di 200 * |predetto - osservato| / (|predetto| + |osservato|).", "Errore percentuale simmetrico", "Limitato a 0-200%, ma puo essere severo vicino a zero."),
            ("RMAE", "100 * MAE / abs(media osservata).", "Confronti percentuali tra variabili", "Instabile se la media osservata e vicina a zero."),
            ("CVRMSE", "100 * RMSE / abs(media osservata).", "Confronti percentuali tra variabili", "Instabile se la media osservata e vicina a zero."),
            ("tp", "True positives: presenze correttamente previste.", "Specie / binario", "Dipende dalla soglia scelta."),
            ("fp", "False positives: presenze previste ma non osservate.", "Specie / binario", "Troppi fp indicano sovrastima delle presenze."),
            ("fn", "False negatives: presenze osservate ma non previste.", "Specie / binario", "Troppi fn indicano omissione del segnale biologico."),
            ("F1", "2TP / (2TP + FP + FN).", "Specie / binario", "Bilancia precision e recall ma ignora TN."),
            ("Jaccard", "TP / (TP + FP + FN).", "Specie / binario", "Piu severo della Sorensen sui mismatch."),
            ("Sorensen", "2TP / (2TP + FP + FN).", "Specie / binario", "Coincide con F1 nel caso binario."),
        ],
        columns=["metric", "meaning", "best_for", "caution"],
    )


def technical_notes(manifest: dict[str, Any]) -> pd.DataFrame:
    precipitation = manifest.get("precipitation_audit") or PRECIPITATION_AUDIT
    return pd.DataFrame(
        [
            (
                "precipitation_unit",
                precipitation.get("display_unit") or precipitation.get("precipitation_unit_it") or "mm/mese",
                "Precipitazione mostrata come totale mensile.",
            ),
            (
                "precipitation_conversion",
                precipitation.get("conversion") or precipitation.get("precipitation_conversion") or "raw_m * 1000",
                "La stessa conversione viene applicata a predicted e observed.",
            ),
            (
                "precipitation_raw_units",
                precipitation.get("units"),
                "Attributo NetCDF letto da climate.tp quando disponibile.",
            ),
            (
                "ndvi_metric_validity",
                "observed_NDVI_native != 0",
                "Per vegetation.NDVI le celle observed=0 sono escluse da MAE, RMSE, bias, correlation, WAPE e sMAPE.",
            ),
        ],
        columns=["topic", "value", "note"],
    )


def coverage_from_metrics(metrics: pd.DataFrame) -> pd.DataFrame:
    if metrics.empty:
        return pd.DataFrame(columns=["year", "n_months", "months"])
    run_index = metrics[["forecast_month", "year"]].drop_duplicates()
    rows = []
    for year, frame in run_index.groupby("year"):
        months = sorted(pd.to_datetime(frame["forecast_month"]).dt.strftime("%Y-%m").unique().tolist())
        rows.append(
            {
                "year": int(year),
                "n_months": int(len(months)),
                "months": ", ".join(months),
            }
        )
    return pd.DataFrame(rows).sort_values("year").reset_index(drop=True)


def filter_metrics(frame: pd.DataFrame, predicate: Callable[[pd.DataFrame], pd.Series]) -> pd.DataFrame:
    output = frame.loc[predicate(frame)].copy()
    return output.reset_index(drop=True)


def filter_curated_feature_metrics(frame: pd.DataFrame, feature: dict[str, Any]) -> pd.DataFrame:
    output = frame[
        (frame["group"] == feature["group"])
        & (frame["variable"] == feature["variable"])
    ].copy()
    if "level_index" in feature:
        output = output[output["level_index"] == feature["level_index"]].copy()
    return output.reset_index(drop=True)


def ensure_area_analysis_scaffold(runs_root: Path) -> dict[str, str]:
    area_root = default_area_analysis_root(runs_root)
    area_root.mkdir(parents=True, exist_ok=True)
    schema_path = area_root / "_schema.json"
    schema_payload = {
        "status": "reserved_for_future_frontend_and_analysis_pipeline",
        "observed_monthly_wide_columns": [
            "month",
            "t2m_mean_area",
            "tp_mean_area",
            "NDVI_mean_area",
            "swvl1_mean_area",
            "swvl2_mean_area",
            "q_mean_area",
            "Forest_mean_area",
            "Arable_mean_area",
            "Cropland_mean_area",
            "avg_snlwrf_mean_area",
            "stl1_mean_area",
            "stl2_mean_area",
        ],
        "forecast_monthly_metrics_long_columns": [
            "month",
            "feature_key",
            "group",
            "variable",
            "level_index",
            "unit",
            "predicted_mean",
            "observed_mean",
            "mae",
            "rmse",
            "correlation",
            "wape_pct",
            "smape_pct",
            "rmae_pct",
            "cvrmse_pct",
            "valid_cell_count",
        ],
        "notes": [
            "Scaffold preparato per analisi area-level e futuro collegamento frontend.",
            "Nessuna modifica frontend viene applicata in questa fase.",
        ],
    }
    schema_path.write_text(json.dumps(schema_payload, indent=2, ensure_ascii=False), encoding="utf-8")
    return {
        "root": str(area_root),
        "schema": str(schema_path),
    }


def build_native_output_index(run_dirs: list[Path], runs_root: Path) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    output_root = output_root_from_runs_root(runs_root)
    previsioni_root = output_root / DEFAULT_PREVISIONI_DIRNAME
    for run_dir in run_dirs:
        manifest_path = run_dir / "forecast_native_manifest.json"
        manifest = load_json(manifest_path)
        forecast_month_ts = safe_timestamp(manifest.get("forecast_month"))
        month_label = forecast_month_ts.strftime("%Y-%m") if forecast_month_ts is not None else ""
        previsioni_month_root = previsioni_root / month_label if month_label else None
        selected_group = str((manifest.get("lab_export") or {}).get("group") or "")
        selected_variable = str((manifest.get("lab_export") or {}).get("variable") or "")
        selected_plot_dir = previsioni_month_root / "plot" / selected_variable if previsioni_month_root and selected_variable else None
        selected_matrix_path = (
            previsioni_month_root / "cell_matrix" / f"{selected_variable}_cell_matrix.xlsx"
            if previsioni_month_root and selected_variable
            else None
        )
        species_matrix_path = (
            previsioni_month_root / "species" / "species_cell_matrix.xlsx"
            if previsioni_month_root
            else None
        )
        native_full_xlsx = run_dir / "exports" / "bioanalyst_native_full_output.xlsx"
        native_group_csv_dir = run_dir / "exports" / "bioanalyst_native_full_output_group_csv"
        native_group_xlsx_dir = run_dir / "exports" / "bioanalyst_native_full_output_group_xlsx"
        rows.append(
            {
                "run_dir": str(run_dir),
                "label": run_label(manifest, run_dir),
                "forecast_month": manifest.get("forecast_month"),
                "manifest_path": str(manifest_path),
                "native_prediction_original": manifest.get("native_prediction_original") or str(run_dir / "native_prediction_original.pt"),
                "native_target_original": manifest.get("native_target_original") or str(run_dir / "native_target_original.pt"),
                "native_full_xlsx": str(native_full_xlsx) if native_full_xlsx.exists() else None,
                "native_group_csv_dir": str(native_group_csv_dir) if native_group_csv_dir.exists() else None,
                "native_group_xlsx_dir": str(native_group_xlsx_dir) if native_group_xlsx_dir.exists() else None,
                "selected_group": selected_group,
                "selected_variable": selected_variable,
                "previsioni_month_dir": str(previsioni_month_root) if previsioni_month_root else None,
                "selected_plot_dir": str(selected_plot_dir) if selected_plot_dir and selected_plot_dir.exists() else None,
                "selected_matrix_xlsx": str(selected_matrix_path) if selected_matrix_path and selected_matrix_path.exists() else None,
                "reliable_plot_root": str(previsioni_month_root / "plot") if previsioni_month_root and (previsioni_month_root / "plot").exists() else None,
                "reliable_cell_matrix_root": str(previsioni_month_root / "cell_matrix") if previsioni_month_root and (previsioni_month_root / "cell_matrix").exists() else None,
                "species_matrix_xlsx": str(species_matrix_path) if species_matrix_path and species_matrix_path.exists() else None,
            }
        )
    return pd.DataFrame(rows)


CHART_VARIABLE_SPECS = CURATED_BIOMAP_FEATURES


def curated_metric_frame(metrics: pd.DataFrame, *, feature: dict[str, Any]) -> pd.DataFrame:
    frame = metrics[
        (metrics["group"] == feature["group"])
        & (metrics["variable"] == feature["variable"])
    ].copy()
    if "level_index" in feature:
        frame = frame[frame["level_index"] == feature["level_index"]].copy()
    if frame.empty:
        return frame
    frame["forecast_month_dt"] = pd.to_datetime(frame["forecast_month"], errors="coerce")
    numeric_columns = [
        "predicted_mean",
        "observed_mean",
        "mae",
        "rmse",
        "correlation",
        "wape_pct",
        "smape_pct",
        "rmae_pct",
        "cvrmse_pct",
    ]
    aggregated = (
        frame.groupby("forecast_month_dt", dropna=False)[numeric_columns]
        .mean(numeric_only=True)
        .reset_index()
        .sort_values("forecast_month_dt")
    )
    aggregated["forecast_month"] = aggregated["forecast_month_dt"].dt.strftime("%Y-%m")
    aggregated["year"] = aggregated["forecast_month_dt"].dt.year
    return aggregated


def save_multi_line_chart(
    frame: pd.DataFrame,
    *,
    output_path: Path,
    title: str,
    y_columns: list[str],
    ylabel: str,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    x_labels = frame["forecast_month"].tolist()
    plt.figure(figsize=(11, 5.5))
    for column in y_columns:
        plt.plot(x_labels, frame[column], marker="o", linewidth=2, label=column)
    plt.title(title)
    plt.xlabel("Forecast month")
    plt.ylabel(ylabel)
    plt.grid(True, alpha=0.25)
    plt.legend()
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()
    plt.savefig(output_path, dpi=180)
    plt.close()


def save_single_metric_chart(
    frame: pd.DataFrame,
    *,
    output_path: Path,
    title: str,
    y_column: str,
    ylabel: str,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    x_labels = frame["forecast_month"].tolist()
    plt.figure(figsize=(11, 5.5))
    plt.plot(x_labels, frame[y_column], marker="o", linewidth=2, color="#2E5EAA")
    plt.title(title)
    plt.xlabel("Forecast month")
    plt.ylabel(ylabel)
    plt.grid(True, alpha=0.25)
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()
    plt.savefig(output_path, dpi=180)
    plt.close()


def save_yearly_summary_chart(frame: pd.DataFrame, *, output_path: Path, title: str) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    yearly = (
        frame.groupby("year", dropna=False)[
            [
                "predicted_mean",
                "observed_mean",
                "mae",
                "rmse",
                "wape_pct",
                "smape_pct",
                "correlation",
                "rmae_pct",
                "cvrmse_pct",
            ]
        ]
        .mean(numeric_only=True)
        .reset_index()
        .sort_values("year")
    )
    if yearly.empty:
        return
    x_labels = [str(int(year)) for year in yearly["year"].tolist()]
    fig, axes = plt.subplots(3, 2, figsize=(13, 12))
    fig.suptitle(title)

    axes[0, 0].plot(x_labels, yearly["predicted_mean"], marker="o", linewidth=2, label="predicted_mean")
    axes[0, 0].plot(x_labels, yearly["observed_mean"], marker="o", linewidth=2, label="observed_mean")
    axes[0, 0].set_title("Observed vs predicted")
    axes[0, 0].grid(True, alpha=0.25)
    axes[0, 0].legend()

    axes[0, 1].plot(x_labels, yearly["mae"], marker="o", linewidth=2, label="MAE")
    axes[0, 1].plot(x_labels, yearly["rmse"], marker="o", linewidth=2, label="RMSE")
    axes[0, 1].set_title("MAE and RMSE")
    axes[0, 1].grid(True, alpha=0.25)
    axes[0, 1].legend()

    axes[1, 0].plot(x_labels, yearly["wape_pct"], marker="o", linewidth=2, label="WAPE")
    axes[1, 0].plot(x_labels, yearly["smape_pct"], marker="o", linewidth=2, label="sMAPE")
    axes[1, 0].set_title("WAPE and sMAPE")
    axes[1, 0].grid(True, alpha=0.25)
    axes[1, 0].legend()

    axes[1, 1].plot(x_labels, yearly["correlation"], marker="o", linewidth=2, color="#2E5EAA")
    axes[1, 1].set_title("Correlation")
    axes[1, 1].grid(True, alpha=0.25)

    axes[2, 0].plot(x_labels, yearly["rmae_pct"], marker="o", linewidth=2, label="RMAE")
    axes[2, 0].plot(x_labels, yearly["cvrmse_pct"], marker="o", linewidth=2, label="CVRMSE")
    axes[2, 0].set_title("RMAE and CVRMSE")
    axes[2, 0].grid(True, alpha=0.25)
    axes[2, 0].legend()

    axes[2, 1].axis("off")

    for axis in axes.flat:
        axis.tick_params(axis="x", rotation=45)
    fig.tight_layout()
    fig.savefig(output_path, dpi=180)
    plt.close(fig)


def generate_feature_chart_artifacts(metrics: pd.DataFrame, *, chart_root: Path) -> dict[str, dict[str, str]]:
    outputs: dict[str, dict[str, str]] = {}
    for feature in CHART_VARIABLE_SPECS:
        frame = curated_metric_frame(metrics, feature=feature)
        if frame.empty:
            continue
        variable = str(feature["variable"])
        folder_name = str(feature["chart_folder"])
        variable_root = chart_root / folder_name
        observed_vs_predicted_path = variable_root / "observed_vs_predicted_by_month.png"
        mae_rmse_path = variable_root / "mae_rmse_by_month.png"
        wape_smape_path = variable_root / "wape_smape_by_month.png"
        correlation_path = variable_root / "correlation_by_month.png"
        yearly_summary_path = variable_root / "yearly_summary.png"
        rmae_cvrmse_path = variable_root / "rmae_cvrmse_by_month.png"

        save_multi_line_chart(
            frame,
            output_path=observed_vs_predicted_path,
            title=f"{variable}: observed vs predicted by month",
            y_columns=["observed_mean", "predicted_mean"],
            ylabel="Mean value",
        )
        save_multi_line_chart(
            frame,
            output_path=mae_rmse_path,
            title=f"{variable}: MAE and RMSE by month",
            y_columns=["mae", "rmse"],
            ylabel="Error",
        )
        save_multi_line_chart(
            frame,
            output_path=wape_smape_path,
            title=f"{variable}: WAPE and sMAPE by month",
            y_columns=["wape_pct", "smape_pct"],
            ylabel="Percent",
        )
        save_single_metric_chart(
            frame,
            output_path=correlation_path,
            title=f"{variable}: correlation by month",
            y_column="correlation",
            ylabel="Correlation",
        )
        save_multi_line_chart(
            frame,
            output_path=rmae_cvrmse_path,
            title=f"{variable}: RMAE and CVRMSE by month",
            y_columns=["rmae_pct", "cvrmse_pct"],
            ylabel="Percent",
        )
        save_yearly_summary_chart(
            frame,
            output_path=yearly_summary_path,
            title=f"{variable}: yearly summary",
        )
        outputs[folder_name] = {
            "observed_vs_predicted_by_month": str(observed_vs_predicted_path),
            "mae_rmse_by_month": str(mae_rmse_path),
            "wape_smape_by_month": str(wape_smape_path),
            "correlation_by_month": str(correlation_path),
            "rmae_cvrmse_by_month": str(rmae_cvrmse_path),
            "yearly_summary": str(yearly_summary_path),
        }
    return outputs


def write_workbook(path: Path, sheets: OrderedDict[str, pd.DataFrame]) -> Path:
    from openpyxl import Workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Font, PatternFill

    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = path.with_name(f".{path.stem}.tmp.xlsx")
    workbook = Workbook(write_only=True)
    header_fill = PatternFill("solid", fgColor="17324D")
    header_font = Font(color="FFFFFF", bold=True)
    total_sheets = len(sheets)

    for sheet_index, (sheet_name, frame) in enumerate(sheets.items(), start=1):
        print(
            f"[workbook] Scrivo foglio {sheet_index}/{total_sheets}: {sheet_name} "
            f"({len(frame):,} righe x {len(frame.columns):,} colonne)",
            flush=True,
        )
        worksheet = workbook.create_sheet(title=sheet_name[:31])
        header_row = []
        for column_name in frame.columns:
            cell = WriteOnlyCell(worksheet, value=column_name)
            cell.fill = header_fill
            cell.font = header_font
            header_row.append(cell)
        worksheet.append(header_row)

        for row_index, row in enumerate(frame.itertuples(index=False, name=None), start=1):
            values = []
            for value in row:
                if isinstance(value, pd.Timestamp):
                    values.append(value.strftime("%Y-%m-%d"))
                elif pd.isna(value):
                    values.append(None)
                else:
                    values.append(value)
            worksheet.append(values)
            if row_index % 10000 == 0:
                print(f"[workbook]   {sheet_name}: {row_index:,}/{len(frame):,} righe", flush=True)

    workbook.save(temp_path)
    try:
        temp_path.replace(path)
        return path
    except PermissionError:
        fallback = path.with_name(f"{path.stem}_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        temp_path.replace(fallback)
        return fallback


def update_biomap_final_workbook(
    *,
    current_run_dir: Path | None = None,
    prediction_batch: Any | None = None,
    observed_batch: Any | None = None,
    manifest: dict[str, Any] | None = None,
    runs_root: Path = DEFAULT_MODEL_FORECAST_ROOT,
    workbook_path: Path | None = None,
) -> dict[str, Any]:
    runs_root = infer_runs_root(current_run_dir, workbook_path, runs_root)
    workbook_path = workbook_path or default_workbook_path(runs_root)
    chart_root = default_chart_root(runs_root)
    area_analysis_scaffold = ensure_area_analysis_scaffold(runs_root)
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    chart_root.mkdir(parents=True, exist_ok=True)

    print(f"[workbook] Scansiono run storici in: {runs_root}", flush=True)
    run_dirs = iter_run_dirs(runs_root)
    if current_run_dir is not None and current_run_dir not in run_dirs:
        run_dirs.append(current_run_dir)
        run_dirs.sort(
            key=lambda path: (
                safe_timestamp(load_json(path / "forecast_native_manifest.json").get("forecast_month")) or pd.Timestamp.max,
                path.name.lower(),
            )
        )
    if not run_dirs:
        raise RuntimeError(f"Nessun run valido trovato sotto {runs_root}")

    if current_run_dir is None:
        current_run_dir = run_dirs[-1]
    if manifest is None:
        manifest = load_json(current_run_dir / "forecast_native_manifest.json")
    if prediction_batch is None:
        prediction_batch = ant.load_batch(current_run_dir / "native_prediction_original.pt")
    if observed_batch is None:
        observed_batch = ant.load_batch(current_run_dir / "native_target_original.pt")

    cached_batches = {current_run_dir: (prediction_batch, observed_batch, manifest)}

    print("[workbook] Calcolo metriche cumulative", flush=True)
    all_metrics = load_all_metrics(run_dirs, current_batches=cached_batches)
    print("[workbook] Calcolo metriche specie", flush=True)
    species_metrics = compute_species_summary(run_dirs, current_batches=cached_batches)
    print("[workbook] Costruisco fogli finali", flush=True)
    dashboard = build_dashboard(all_metrics)
    indicator_map = build_indicator_map(dashboard)
    coverage = coverage_from_metrics(all_metrics)
    species_index = build_species_index(all_metrics)
    native_output_index = build_native_output_index(run_dirs, runs_root)

    sheets: OrderedDict[str, pd.DataFrame] = OrderedDict()
    sheets["Dashboard_BIOMAP"] = dashboard
    sheets["BIOMAP_Indicator_Map"] = indicator_map
    sheets["Metric_Guide"] = metric_guide()
    sheets["Technical_Notes"] = technical_notes(manifest)
    sheets["Coverage"] = coverage
    sheets["Native_Output_Index"] = native_output_index
    for feature in CURATED_BIOMAP_FEATURES:
        sheets[str(feature["sheet_name"])] = filter_curated_feature_metrics(all_metrics, feature)
    sheets["Species_Biodiversity"] = species_metrics
    sheets["Species_Index"] = species_index
    sheets["Agriculture_All"] = filter_metrics(all_metrics, lambda frame: frame["group"] == "agriculture")
    sheets["Climate_Other"] = filter_metrics(all_metrics, lambda frame: (frame["group"] == "climate") & (~frame["variable"].isin(["t2m", "tp"])))
    sheets["Edaphic_Land"] = filter_metrics(all_metrics, lambda frame: frame["group"].isin(["edaphic", "land"]))
    sheets["Surface"] = filter_metrics(all_metrics, lambda frame: frame["group"] == "surface")
    sheets["Atmospheric_All"] = filter_metrics(all_metrics, lambda frame: frame["group"] == "atmospheric")
    sheets["Misc_Redlist"] = filter_metrics(all_metrics, lambda frame: frame["group"].isin(["misc", "redlist"]))
    sheets["Species_Native_By_ID"] = filter_metrics(all_metrics, lambda frame: frame["group"] == "species")
    sheets["All_Variables"] = all_metrics

    print("[workbook] Genero chart cumulativi", flush=True)
    chart_outputs = generate_feature_chart_artifacts(all_metrics, chart_root=chart_root)
    print("[workbook] Scrittura workbook Excel", flush=True)
    written_workbook = write_workbook(workbook_path, sheets)
    print(f"[workbook] Workbook scritto: {written_workbook}", flush=True)
    return {
        "workbook": str(written_workbook),
        "chart_root": str(chart_root),
        "area_analysis_root": area_analysis_scaffold["root"],
        "area_analysis_schema": area_analysis_scaffold["schema"],
        "charts": chart_outputs,
        "runs_used": len(run_dirs),
        "latest_run": str(current_run_dir),
        "sheet_count": len(sheets),
    }


def main() -> None:
    args = build_parser().parse_args()
    result = update_biomap_final_workbook(
        current_run_dir=args.run_dir.resolve() if args.run_dir else None,
        runs_root=args.runs_root.resolve(),
        workbook_path=args.workbook_path.resolve() if args.workbook_path else None,
    )
    print(json.dumps(result, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
