#!/usr/bin/env python3
"""Esegue un caso BioAnalyst native e prepara output leggibili da tesi.

Un solo comando produce:
- forecast native one-step;
- PNG prediction, observed e difference per la variabile scelta;
- Excel/CSV con matrice cella-per-cella e riepilogo errori.
"""

from __future__ import annotations

import json
import math
import os
import inspect
import shutil
from collections import OrderedDict
from pathlib import Path
from typing import Any

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import torch
import torch.nn.functional as F
from omegaconf import OmegaConf

from bioanalyst_model_utils import build_selection_parser
from bioanalyst_model_utils import resolve_city_bounds
from biomap_final_workbook import DEFAULT_WORKBOOK_NAME, update_biomap_final_workbook
from bioanalyst_native_utils import (
    build_native_run_context,
    build_native_runtime,
    ensure_bfm_repo_on_path,
    prepare_native_forecast_environment,
    prepare_native_saved_windows,
    run_native_one_step,
    save_native_one_step_artifacts,
)
from plot_native_maps import add_selection_box, get_batch_variable_map, get_grid_coordinates
from spatial_alignment import (
    align_prediction_map,
    build_alignment_diagnostic_frame,
    build_spatial_alignment_metadata,
    plot_origin_for_latitudes,
    prediction_latitude_flip_enabled,
    prediction_longitude_flip_enabled,
)


RELIABLE_PLOT_FEATURES = [
    {
        "folder": "temperature",
        "file_prefix": "t2m",
        "group": "climate",
        "variable": "t2m",
        "cmap": "coolwarm",
    },
    {
        "folder": "ndvi",
        "file_prefix": "ndvi",
        "group": "vegetation",
        "variable": "NDVI",
        "cmap": "YlGn",
    },
    {
        "folder": "swvl1",
        "file_prefix": "swvl1",
        "group": "edaphic",
        "variable": "swvl1",
        "cmap": "Blues",
    },
    {
        "folder": "swvl2",
        "file_prefix": "swvl2",
        "group": "edaphic",
        "variable": "swvl2",
        "cmap": "Blues",
    },
    {
        "folder": "cropland",
        "file_prefix": "cropland",
        "group": "agriculture",
        "variable": "Cropland",
        "cmap": "YlOrBr",
    },
    {
        "folder": "precipitation",
        "file_prefix": "tp",
        "group": "climate",
        "variable": "tp",
        "cmap": "PuBu",
    },
]


def previsioni_root_for_run(run_dir: Path) -> Path:
    return run_dir.parent / "previsioni"


def previsioni_month_dir(run_dir: Path, manifest: dict[str, Any]) -> Path:
    forecast_month = str(manifest.get("forecast_month") or "").strip()
    if forecast_month:
        month_label = pd.Timestamp(forecast_month).strftime("%Y-%m")
    else:
        month_label = run_dir.name
    return previsioni_root_for_run(run_dir) / month_label


def build_parser():
    parser = build_selection_parser("Run unico per forecast BioAnalyst native con mappe ed Excel di validazione.")
    parser.add_argument("--checkpoint", choices=["small", "large"], default="small")
    parser.add_argument("--device", choices=["auto", "cpu", "cuda", "mps"], default="cuda")
    parser.add_argument(
        "--input-mode",
        choices=["clean", "all"],
        default="all",
        help="clean usa solo input core; all aggiunge vegetation/agriculture/forest quando disponibili.",
    )
    parser.add_argument("--group", default="climate", help="Gruppo nativo da esportare.")
    parser.add_argument("--variable", default="t2m", help="Variabile nativa da esportare.")
    parser.add_argument(
        "--fast-smoke-test",
        action="store_true",
        help="Usa placeholder atmosferici a zero per verificare velocemente la pipeline.",
    )
    parser.add_argument(
        "--no-compare-observed",
        action="store_true",
        help="Non salva il target observed. Se attivo non crea difference e metriche di errore.",
    )
    parser.add_argument("--amp-bf16", action="store_true", help="Usa autocast bfloat16 su CUDA.")
    parser.add_argument(
        "--attention-chunk-size",
        type=int,
        default=None,
        help="Se impostato, processa l'attenzione CUDA a blocchi per ridurre VRAM. Esempio: 32 o 16.",
    )
    parser.add_argument(
        "--summary-cities",
        default="",
        help="Città extra da riassumere nello stesso Excel, separate da virgola. Esempio: Milan,Madrid,Paris.",
    )
    parser.add_argument(
        "--summary-half-window-deg",
        type=float,
        default=None,
        help="Semiampiezza area per le città extra. Se omesso usa --half-window-deg.",
    )
    parser.add_argument(
        "--history-xlsx",
        type=Path,
        default=None,
        help="Workbook cumulativo da aggiornare a ogni run. Se omesso usa un file separato per checkpoint.",
    )
    parser.add_argument(
        "--no-history",
        action="store_true",
        help="Non aggiorna il workbook cumulativo.",
    )
    parser.add_argument(
        "--export-native-full",
        action="store_true",
        help="Esporta anche un workbook leggibile dell'output .pt nativo completo.",
    )
    parser.add_argument(
        "--export-native-group-csvs",
        action="store_true",
        help="Con --export-native-full, esporta anche un CSV per ogni gruppo nativo con tutte le variabili.",
    )
    parser.add_argument(
        "--biomap-final-workbook",
        type=Path,
        default=None,
        help="Workbook finale BIOMAP da aggiornare automaticamente. Se omesso usa il percorso canonico.",
    )
    parser.add_argument(
        "--no-biomap-final-workbook",
        action="store_true",
        help="Non aggiorna il workbook finale BIOMAP a fine run.",
    )
    parser.add_argument(
        "--matrix-export-format",
        choices=["excel", "csv", "both"],
        default="excel",
        help="Formato di export per i cell matrix delle feature principali e delle specie.",
    )
    return parser


def install_chunked_cuda_attention(chunk_size: int) -> None:
    """Patch runtime: evita attention matrix giganti senza modificare file esterni."""
    ensure_bfm_repo_on_path()
    from bfm_model.perceiver_components import helpers

    def chunked_forward(self, x: torch.Tensor, context: torch.Tensor = None, mask: torch.Tensor = None, is_causal: bool = None):
        if context is None:
            context = x
        is_causal_value = self.is_causal if is_causal is None else is_causal

        bsz, seq_len_q, _ = x.shape
        seq_len_kv = context.size(1)

        q = self.q_proj(x)
        k = self.k_proj(context)
        v = self.v_proj(context)

        q = q.reshape(bsz, seq_len_q, self.n_q_heads, self.head_dim).transpose(1, 2)
        k = k.reshape(bsz, seq_len_kv, self.n_kv_heads, self.head_dim).transpose(1, 2)
        v = v.reshape(bsz, seq_len_kv, self.n_kv_heads, self.head_dim).transpose(1, 2)

        final_attn_mask = mask.unsqueeze(1) if mask is not None else None
        use_chunked_attention = q.device.type == "cuda" and seq_len_q > chunk_size

        if use_chunked_attention:
            if self.n_kv_heads < self.n_q_heads:
                repeat_factor = self.n_q_heads // self.n_kv_heads
                k = k.repeat_interleave(repeat_factor, dim=1)
                v = v.repeat_interleave(repeat_factor, dim=1)

            chunks = []
            for start_idx in range(0, seq_len_q, chunk_size):
                end_idx = min(start_idx + chunk_size, seq_len_q)
                attn_mask_chunk = None
                if final_attn_mask is not None:
                    attn_mask_chunk = final_attn_mask[:, :, start_idx:end_idx, :]
                chunks.append(
                    F.scaled_dot_product_attention(
                        q[:, :, start_idx:end_idx, :],
                        k,
                        v,
                        attn_mask=attn_mask_chunk,
                        dropout_p=self.dropout_p if self.training else 0.0,
                        is_causal=is_causal_value,
                        enable_gqa=False,
                    )
                )
            out = torch.cat(chunks, dim=2)
        else:
            out = F.scaled_dot_product_attention(
                q,
                k,
                v,
                attn_mask=final_attn_mask,
                dropout_p=self.dropout_p if self.training else 0.0,
                is_causal=is_causal_value,
                enable_gqa=True,
            )

        out = out.transpose(1, 2).reshape(bsz, seq_len_q, self.n_q_heads * self.head_dim)
        return self.out_proj(out)

    helpers.BuiltinGQAAttention.forward = chunked_forward
    print(f"[cuda] Attention chunking runtime attivo: chunk_size={chunk_size}", flush=True)


def plot_map(
    output_path: Path,
    map_values: np.ndarray,
    latitudes: np.ndarray,
    longitudes: np.ndarray,
    title: str,
    unit: str,
    bounds: dict[str, float],
    cmap: str = "coolwarm",
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig, ax = plt.subplots(figsize=(11, 6))
    image = ax.imshow(
        map_values,
        origin=plot_origin_for_latitudes(latitudes),
        extent=[float(longitudes.min()), float(longitudes.max()), float(latitudes.min()), float(latitudes.max())],
        aspect="auto",
        cmap=cmap,
    )
    add_selection_box(ax, bounds)
    ax.set_title(title)
    ax.set_xlabel("Longitude")
    ax.set_ylabel("Latitude")
    colorbar = fig.colorbar(image, ax=ax)
    colorbar.set_label(unit)
    fig.tight_layout()
    fig.savefig(output_path, dpi=200)
    plt.close(fig)


def write_excel_friendly_csv(path: Path, frame: pd.DataFrame) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    frame.to_csv(path, index=False, encoding="utf-8-sig", sep=";", decimal=",")


def export_mode_wants_excel(mode: str) -> bool:
    return mode in {"excel", "both"}


def export_mode_wants_csv(mode: str) -> bool:
    return mode in {"csv", "both"}


def write_streaming_workbook(path: Path, sheets: OrderedDict[str, pd.DataFrame]) -> Path:
    from openpyxl import Workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Font, PatternFill

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook(write_only=True)
    header_fill = PatternFill("solid", fgColor="17324D")
    header_font = Font(color="FFFFFF", bold=True)
    total_sheets = len(sheets)

    for sheet_index, (sheet_name, frame) in enumerate(sheets.items(), start=1):
        print(
            f"[excel] Scrivo foglio {sheet_index}/{total_sheets}: {sheet_name} "
            f"({len(frame):,} righe x {len(frame.columns):,} colonne)",
            flush=True,
        )
        worksheet = workbook.create_sheet(title=sheet_name[:31])
        header_row = []
        for column_name in frame.columns:
            cell = WriteOnlyCell(worksheet, value=column_name)
            cell.fill = header_fill
            cell.font = header_font
            header_row.append(cell)
        worksheet.append(header_row)

        for row_index, row in enumerate(frame.itertuples(index=False, name=None), start=1):
            values = []
            for value in row:
                if isinstance(value, pd.Timestamp):
                    values.append(value.strftime("%Y-%m-%d"))
                elif pd.isna(value):
                    values.append(None)
                else:
                    values.append(value)
            worksheet.append(values)
            if row_index % 10000 == 0:
                print(f"[excel]   {sheet_name}: {row_index:,}/{len(frame):,} righe", flush=True)

    temp_path = path.with_name(f".{path.stem}.tmp.xlsx")
    workbook.save(temp_path)
    try:
        temp_path.replace(path)
        return path
    except PermissionError:
        fallback = path.with_name(f"{path.stem}_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        temp_path.replace(fallback)
        return fallback


def export_reliable_feature_plots(
    run_dir: Path,
    manifest: dict[str, Any],
    predicted_batch: Any,
    observed_batch: Any | None,
) -> dict[str, Any]:
    """Salva prediction/observed/difference PNG per le feature BIOMAP piu utili."""
    plot_root = run_dir / "plots" / "reliable_features"
    predicted_batch = move_tensors_to_cpu(predicted_batch)
    observed_batch = move_tensors_to_cpu(observed_batch) if observed_batch is not None else None

    latitudes, longitudes = get_grid_coordinates(predicted_batch)
    bounds = manifest["bounds"]
    latitude_flip = prediction_latitude_flip_enabled(manifest)
    longitude_flip = prediction_longitude_flip_enabled(manifest)
    outputs: dict[str, Any] = {
        "root": str(plot_root),
        "features": {},
        "missing_features": [],
    }

    for feature in RELIABLE_PLOT_FEATURES:
        group = feature["group"]
        variable = feature["variable"]
        file_prefix = feature["file_prefix"]
        feature_dir = plot_root / feature["folder"]
        try:
            predicted_map_raw, unit = get_batch_variable_map(predicted_batch, group_name=group, variable_name=variable)
        except SystemExit as exc:
            outputs["missing_features"].append({"group": group, "variable": variable, "reason": str(exc)})
            continue

        predicted_map = align_prediction_map(
            predicted_map_raw,
            latitude_flip=latitude_flip,
            longitude_flip=longitude_flip,
        )
        prediction_png = feature_dir / f"{file_prefix}_prediction.png"
        plot_map(
            prediction_png,
            predicted_map,
            latitudes,
            longitudes,
            f"{group}.{variable} prediction",
            unit,
            bounds,
            cmap=feature["cmap"],
        )

        feature_outputs = {"prediction_png": str(prediction_png)}
        if observed_batch is not None:
            try:
                observed_map, _ = get_batch_variable_map(observed_batch, group_name=group, variable_name=variable)
            except SystemExit as exc:
                outputs["missing_features"].append({"group": group, "variable": variable, "reason": str(exc)})
            else:
                observed_png = feature_dir / f"{file_prefix}_observed.png"
                difference_png = feature_dir / f"{file_prefix}_difference.png"
                plot_map(
                    observed_png,
                    observed_map,
                    latitudes,
                    longitudes,
                    f"{group}.{variable} observed",
                    unit,
                    bounds,
                    cmap=feature["cmap"],
                )
                plot_map(
                    difference_png,
                    predicted_map - observed_map,
                    latitudes,
                    longitudes,
                    f"{group}.{variable} prediction - observed",
                    unit,
                    bounds,
                    cmap="coolwarm",
                )
                feature_outputs["observed_png"] = str(observed_png)
                feature_outputs["difference_png"] = str(difference_png)

        outputs["features"][file_prefix] = feature_outputs

    return outputs


def build_cell_frame(
    *,
    latitudes: np.ndarray,
    longitudes: np.ndarray,
    predicted_map: np.ndarray,
    observed_map: np.ndarray | None,
    bounds: dict[str, float],
    value_label: str,
) -> pd.DataFrame:
    rows = []
    for lat_index, lat in enumerate(latitudes):
        for lon_index, lon in enumerate(longitudes):
            observed_value = None if observed_map is None else float(observed_map[lat_index, lon_index])
            predicted_value = float(predicted_map[lat_index, lon_index])
            difference = None if observed_value is None else predicted_value - observed_value
            rows.append(
                {
                    "lat": float(lat),
                    "lon": float(lon),
                    f"predicted_{value_label}": predicted_value,
                    f"observed_{value_label}": observed_value,
                    f"difference_{value_label}": difference,
                    f"abs_error_{value_label}": None if difference is None else abs(difference),
                    "inside_selected_area": bool(
                        bounds["min_lat"] <= float(lat) <= bounds["max_lat"]
                        and bounds["min_lon"] <= float(lon) <= bounds["max_lon"]
                    ),
                }
            )
    return pd.DataFrame.from_records(rows)


def compute_binary_summary(predicted_values: np.ndarray, observed_values: np.ndarray, threshold: float = 0.5) -> dict[str, Any]:
    pred_mask = np.asarray(predicted_values) >= threshold
    obs_mask = np.asarray(observed_values) >= threshold
    tp = int(np.logical_and(pred_mask, obs_mask).sum())
    fp = int(np.logical_and(pred_mask, ~obs_mask).sum())
    fn = int(np.logical_and(~pred_mask, obs_mask).sum())
    tn = int(np.logical_and(~pred_mask, ~obs_mask).sum())
    precision = None if tp + fp == 0 else float(tp / (tp + fp))
    recall = None if tp + fn == 0 else float(tp / (tp + fn))
    f1 = None if (2 * tp + fp + fn) == 0 else float((2 * tp) / (2 * tp + fp + fn))
    jaccard = None if (tp + fp + fn) == 0 else float(tp / (tp + fp + fn))
    sorensen = f1
    return {
        "binary_threshold": threshold,
        "tp": tp,
        "fp": fp,
        "fn": fn,
        "tn": tn,
        "precision": precision,
        "recall": recall,
        "f1_score": f1,
        "jaccard_similarity": jaccard,
        "sorensen_similarity": sorensen,
    }


def parse_city_list(value: str) -> list[str]:
    """Legge una lista semplice tipo `Milan,Madrid,Paris`."""
    if not value:
        return []
    normalized = value.replace(";", ",")
    return [item.strip() for item in normalized.split(",") if item.strip()]


def bounds_center(bounds: dict[str, float]) -> tuple[float, float]:
    return (
        float((bounds["min_lat"] + bounds["max_lat"]) / 2.0),
        float((bounds["min_lon"] + bounds["max_lon"]) / 2.0),
    )


def select_area_cells(frame: pd.DataFrame, bounds: dict[str, float]) -> pd.DataFrame:
    return frame[
        (frame["lat"] >= bounds["min_lat"])
        & (frame["lat"] <= bounds["max_lat"])
        & (frame["lon"] >= bounds["min_lon"])
        & (frame["lon"] <= bounds["max_lon"])
    ].copy()


def build_area_specs(args: Any, manifest: dict[str, Any]) -> list[dict[str, Any]]:
    """Prepara l'area principale e le eventuali città extra da riassumere."""
    specs = [
        {
            "area_label": manifest.get("label", "selected_area"),
            "area_kind": manifest.get("selection_mode", "selected_area"),
            "bounds": manifest["bounds"],
            "is_primary_area": True,
        }
    ]

    half_window = args.summary_half_window_deg if args.summary_half_window_deg is not None else args.half_window_deg
    seen = {json.dumps(manifest["bounds"], sort_keys=True)}
    for city_name in parse_city_list(args.summary_cities):
        bounds, resolved_name = resolve_city_bounds(city_name, half_window)
        bounds_key = json.dumps(bounds, sort_keys=True)
        if bounds_key in seen:
            continue
        seen.add(bounds_key)
        specs.append(
            {
                "area_label": resolved_name,
                "area_kind": "city_extra",
                "bounds": bounds,
                "is_primary_area": False,
            }
        )
    return specs


def move_tensors_to_cpu(value: Any) -> Any:
    """Rende sicuri per NumPy/Pandas i batch prodotti su CUDA."""
    if isinstance(value, torch.Tensor):
        return value.detach().cpu()
    if isinstance(value, dict):
        return {key: move_tensors_to_cpu(item) for key, item in value.items()}
    if isinstance(value, list):
        return [move_tensors_to_cpu(item) for item in value]
    if isinstance(value, tuple) and hasattr(value, "_fields"):
        return type(value)(*(move_tensors_to_cpu(item) for item in value))
    if isinstance(value, tuple):
        return tuple(move_tensors_to_cpu(item) for item in value)
    return value


def compute_summary(frame: pd.DataFrame, value_label: str, scope: str) -> dict[str, Any]:
    pred_col = f"predicted_{value_label}"
    obs_col = f"observed_{value_label}"
    diff_col = f"difference_{value_label}"
    abs_col = f"abs_error_{value_label}"

    summary = {
        "scope": scope,
        "cell_count": int(len(frame)),
        "predicted_mean": float(frame[pred_col].mean()),
        "predicted_min": float(frame[pred_col].min()),
        "predicted_max": float(frame[pred_col].max()),
    }
    if obs_col in frame and frame[obs_col].notna().any():
        observed_mean = float(frame[obs_col].mean())
        mae = float(frame[abs_col].mean())
        rmse = float(math.sqrt(float((frame[diff_col] ** 2).mean())))
        bias = float(frame[diff_col].mean())
        summary.update(
            {
                "observed_mean": observed_mean,
                "mae": mae,
                "rmse": rmse,
                "bias": bias,
                "error_pct_on_abs_observed_mean": None
                if abs(observed_mean) < 1e-12
                else float(mae / abs(observed_mean) * 100.0),
            }
        )
    return summary


def export_feature_matrix_bundle(
    *,
    run_dir: Path,
    export_dir: Path,
    workbook_prefix: str,
    area_summary_prefix: str,
    manifest: dict[str, Any],
    latitudes: np.ndarray,
    longitudes: np.ndarray,
    predicted_map: np.ndarray,
    observed_map: np.ndarray | None,
    predicted_map_raw: np.ndarray | None,
    group: str,
    variable: str,
    unit: str,
    area_specs: list[dict[str, Any]],
    matrix_export_format: str = "excel",
) -> tuple[dict[str, str], pd.DataFrame]:
    export_dir.mkdir(parents=True, exist_ok=True)
    bounds = manifest["bounds"]
    latitude_flip = prediction_latitude_flip_enabled(manifest)
    longitude_flip = prediction_longitude_flip_enabled(manifest)
    alignment_diagnostic_frame = None
    if predicted_map_raw is not None and observed_map is not None:
        alignment_diagnostic_frame = build_alignment_diagnostic_frame(predicted_map_raw, observed_map)

    value_label = f"{variable}_{unit.replace('°', '').lower()}"
    frame = build_cell_frame(
        latitudes=latitudes,
        longitudes=longitudes,
        predicted_map=predicted_map,
        observed_map=observed_map,
        bounds=bounds,
        value_label=value_label,
    )
    area_frame = frame[frame["inside_selected_area"]].copy()
    summary_frame = pd.DataFrame.from_records(
        [
            compute_summary(frame, value_label, "full_grid"),
            compute_summary(area_frame, value_label, "selected_area"),
        ]
    )
    summary_frame["prediction_latitude_flip_applied"] = bool(latitude_flip)
    summary_frame["prediction_longitude_flip_applied"] = bool(longitude_flip)
    area_summary_frame = build_area_summary_frame(
        frame=frame,
        value_label=value_label,
        area_specs=area_specs,
        manifest=manifest,
        group=group,
        variable=variable,
        unit=unit,
        run_dir=run_dir,
    )

    outputs: dict[str, str] = {}
    if export_mode_wants_excel(matrix_export_format):
        workbook_path = export_dir / f"{workbook_prefix}_cell_matrix.xlsx"
        with pd.ExcelWriter(workbook_path) as writer:
            summary_frame.to_excel(writer, sheet_name="summary", index=False)
            area_summary_frame.to_excel(writer, sheet_name="area_summary", index=False)
            if alignment_diagnostic_frame is not None:
                alignment_diagnostic_frame.to_excel(writer, sheet_name="alignment_diagnostic", index=False)
            area_frame.to_excel(writer, sheet_name="selected_area", index=False)
            frame.to_excel(writer, sheet_name="full_grid", index=False)
        outputs["matrix_xlsx"] = str(workbook_path)
    if export_mode_wants_csv(matrix_export_format):
        summary_csv_path = export_dir / f"{workbook_prefix}_summary.csv"
        full_csv_path = export_dir / f"{workbook_prefix}_full_grid.csv"
        area_csv_path = export_dir / f"{workbook_prefix}_selected_area.csv"
        area_summary_csv_path = export_dir / f"{workbook_prefix}_area_summary.csv"
        write_excel_friendly_csv(summary_csv_path, summary_frame)
        write_excel_friendly_csv(area_summary_csv_path, area_summary_frame)
        write_excel_friendly_csv(full_csv_path, frame)
        write_excel_friendly_csv(area_csv_path, area_frame)
        outputs["summary_csv"] = str(summary_csv_path)
        outputs["full_grid_csv"] = str(full_csv_path)
        outputs["selected_area_csv"] = str(area_csv_path)
        outputs["area_summary_csv"] = str(area_summary_csv_path)
        if alignment_diagnostic_frame is not None:
            alignment_csv_path = export_dir / f"{workbook_prefix}_alignment_diagnostic.csv"
            write_excel_friendly_csv(alignment_csv_path, alignment_diagnostic_frame)
            outputs["alignment_diagnostic_csv"] = str(alignment_csv_path)
    return outputs, area_summary_frame


def export_reliable_feature_workbooks(
    run_dir: Path,
    manifest: dict[str, Any],
    predicted_batch: Any,
    observed_batch: Any | None,
    area_specs: list[dict[str, Any]],
    export_root: Path | None = None,
    flat_output: bool = False,
    matrix_export_format: str = "excel",
) -> dict[str, Any]:
    export_root = export_root or (run_dir / "exports" / "reliable_feature_workbooks")
    predicted_batch = move_tensors_to_cpu(predicted_batch)
    observed_batch = move_tensors_to_cpu(observed_batch) if observed_batch is not None else None
    latitudes, longitudes = get_grid_coordinates(predicted_batch)
    latitude_flip = prediction_latitude_flip_enabled(manifest)
    longitude_flip = prediction_longitude_flip_enabled(manifest)
    outputs: dict[str, Any] = {"root": str(export_root), "features": {}, "missing_features": []}

    for feature in RELIABLE_PLOT_FEATURES:
        group = feature["group"]
        variable = feature["variable"]
        file_prefix = feature["file_prefix"]
        feature_dir = export_root if flat_output else export_root / feature["folder"]
        try:
            predicted_map_raw, unit = get_batch_variable_map(predicted_batch, group_name=group, variable_name=variable)
        except SystemExit as exc:
            outputs["missing_features"].append({"group": group, "variable": variable, "reason": str(exc)})
            continue
        predicted_map = align_prediction_map(
            predicted_map_raw,
            latitude_flip=latitude_flip,
            longitude_flip=longitude_flip,
        )
        observed_map = None
        if observed_batch is not None:
            try:
                observed_map, _ = get_batch_variable_map(observed_batch, group_name=group, variable_name=variable)
            except SystemExit as exc:
                outputs["missing_features"].append({"group": group, "variable": variable, "reason": str(exc)})
        feature_outputs, _ = export_feature_matrix_bundle(
            run_dir=run_dir,
            export_dir=feature_dir,
            workbook_prefix=file_prefix,
            area_summary_prefix=file_prefix,
            manifest=manifest,
            latitudes=latitudes,
            longitudes=longitudes,
            predicted_map=predicted_map,
            observed_map=observed_map,
            predicted_map_raw=predicted_map_raw,
            group=group,
            variable=variable,
            unit=unit,
            area_specs=area_specs,
            matrix_export_format=matrix_export_format,
        )
        outputs["features"][file_prefix] = feature_outputs
    return outputs


def export_species_matrix_workbook(
    run_dir: Path,
    manifest: dict[str, Any],
    predicted_batch: Any,
    observed_batch: Any | None,
    area_specs: list[dict[str, Any]],
    export_root: Path | None = None,
    flat_output: bool = False,
    matrix_export_format: str = "excel",
) -> dict[str, str] | dict[str, Any]:
    if observed_batch is None:
        return {"reason": "observed batch mancante"}
    export_dir = export_root if flat_output else (export_root or (run_dir / "exports" / "reliable_feature_workbooks")) / "species"
    export_dir.mkdir(parents=True, exist_ok=True)
    predicted_batch = move_tensors_to_cpu(predicted_batch)
    observed_batch = move_tensors_to_cpu(observed_batch)
    latitudes, longitudes = get_grid_coordinates(predicted_batch)
    latitude_flip = prediction_latitude_flip_enabled(manifest)
    longitude_flip = prediction_longitude_flip_enabled(manifest)
    species_group_pred = predicted_batch.species_variables
    species_group_obs = observed_batch.species_variables

    summary_rows: list[dict[str, Any]] = []
    sheets: OrderedDict[str, pd.DataFrame] = OrderedDict()
    index_rows: list[dict[str, Any]] = []

    for species_id in sorted(species_group_pred):
        if species_id not in species_group_obs:
            continue
        predicted_map_raw, unit = get_batch_variable_map(predicted_batch, group_name="species", variable_name=species_id)
        observed_map, _ = get_batch_variable_map(observed_batch, group_name="species", variable_name=species_id)
        predicted_map = align_prediction_map(
            predicted_map_raw,
            latitude_flip=latitude_flip,
            longitude_flip=longitude_flip,
        )
        value_label = f"species_{species_id}"
        frame = build_cell_frame(
            latitudes=latitudes,
            longitudes=longitudes,
            predicted_map=predicted_map,
            observed_map=observed_map,
            bounds=manifest["bounds"],
            value_label=value_label,
        )
        frame["species_id"] = species_id
        frame["unit"] = unit
        frame["forecast_month"] = manifest.get("forecast_month")
        frame["label"] = manifest.get("label")
        frame["group"] = "species"
        frame["variable"] = species_id
        frame["run_dir"] = str(run_dir)
        frame["checkpoint"] = manifest.get("checkpoint_kind")
        frame["input_mode"] = manifest.get("input_mode")
        sheets[f"sp_{species_id}"] = frame

        full_metrics = compute_summary(frame, value_label, "full_grid")
        area_frame = frame[frame["inside_selected_area"]].copy()
        area_metrics = compute_summary(area_frame, value_label, "selected_area")
        binary_full = compute_binary_summary(frame[f"predicted_{value_label}"].to_numpy(), frame[f"observed_{value_label}"].to_numpy())
        binary_area = compute_binary_summary(area_frame[f"predicted_{value_label}"].to_numpy(), area_frame[f"observed_{value_label}"].to_numpy())
        summary_rows.append(
            {
                "forecast_month": manifest.get("forecast_month"),
                "forecast_year": pd.Timestamp(manifest.get("forecast_month")).year if manifest.get("forecast_month") else None,
                "forecast_month_num": pd.Timestamp(manifest.get("forecast_month")).month if manifest.get("forecast_month") else None,
                "label": manifest.get("label"),
                "area_label": manifest.get("label"),
                "species_id": species_id,
                "unit": unit,
                "scope": "full_grid",
                "checkpoint": manifest.get("checkpoint_kind"),
                "input_mode": manifest.get("input_mode"),
                "run_dir": str(run_dir),
                **full_metrics,
                **binary_full,
            }
        )
        summary_rows.append(
            {
                "forecast_month": manifest.get("forecast_month"),
                "forecast_year": pd.Timestamp(manifest.get("forecast_month")).year if manifest.get("forecast_month") else None,
                "forecast_month_num": pd.Timestamp(manifest.get("forecast_month")).month if manifest.get("forecast_month") else None,
                "label": manifest.get("label"),
                "area_label": manifest.get("label"),
                "species_id": species_id,
                "unit": unit,
                "scope": "selected_area",
                "checkpoint": manifest.get("checkpoint_kind"),
                "input_mode": manifest.get("input_mode"),
                "run_dir": str(run_dir),
                **area_metrics,
                **binary_area,
            }
        )
        index_rows.append(
            {
                "species_id": species_id,
                "sheet_name": f"sp_{species_id}",
                "scientific_name": None,
                "common_name": None,
                "note": "ID specie del canale target BioAnalyst",
            }
        )

    summary_frame = pd.DataFrame(summary_rows)
    index_frame = pd.DataFrame(index_rows)
    workbook_sheets: OrderedDict[str, pd.DataFrame] = OrderedDict()
    workbook_sheets["species_summary"] = summary_frame
    workbook_sheets["species_index"] = index_frame
    for name, frame in sheets.items():
        workbook_sheets[name] = frame

    outputs: dict[str, str] = {}
    if export_mode_wants_excel(matrix_export_format):
        workbook_path = write_streaming_workbook(export_dir / "species_cell_matrix.xlsx", workbook_sheets)
        outputs["matrix_xlsx"] = str(workbook_path)
    if export_mode_wants_csv(matrix_export_format):
        write_excel_friendly_csv(export_dir / "species_summary.csv", summary_frame)
        write_excel_friendly_csv(export_dir / "species_index.csv", index_frame)
        for name, frame in sheets.items():
            write_excel_friendly_csv(export_dir / f"{name}.csv", frame)
        outputs["summary_csv"] = str(export_dir / "species_summary.csv")
        outputs["index_csv"] = str(export_dir / "species_index.csv")
    return outputs


def build_area_summary_frame(
    *,
    frame: pd.DataFrame,
    value_label: str,
    area_specs: list[dict[str, Any]],
    manifest: dict[str, Any],
    group: str,
    variable: str,
    unit: str,
    run_dir: Path,
) -> pd.DataFrame:
    rows = []
    spatial_alignment = manifest.get("spatial_alignment", {})
    for spec in area_specs:
        area_frame = select_area_cells(frame, spec["bounds"])
        area_lat, area_lon = bounds_center(spec["bounds"])
        summary = compute_summary(area_frame, value_label, spec["area_label"]) if not area_frame.empty else {
            "scope": spec["area_label"],
            "cell_count": 0,
        }
        error_pct = summary.get("error_pct_on_abs_observed_mean")
        rows.append(
            {
                "run_dir": str(run_dir),
                "label": manifest.get("label"),
                "area_label": spec["area_label"],
                "area_kind": spec["area_kind"],
                "is_primary_area": bool(spec["is_primary_area"]),
                "center_lat": area_lat,
                "center_lon": area_lon,
                "min_lat": spec["bounds"]["min_lat"],
                "max_lat": spec["bounds"]["max_lat"],
                "min_lon": spec["bounds"]["min_lon"],
                "max_lon": spec["bounds"]["max_lon"],
                "input_month_1": manifest.get("input_months", [None, None])[0],
                "input_month_2": manifest.get("input_months", [None, None])[1],
                "forecast_month": manifest.get("forecast_month"),
                "checkpoint_kind": manifest.get("checkpoint_kind"),
                "input_mode": manifest.get("input_mode"),
                "device": manifest.get("device"),
                "prediction_latitude_flip_applied": bool(spatial_alignment.get("prediction_latitude_flip_applied", False)),
                "prediction_longitude_flip_applied": bool(spatial_alignment.get("prediction_longitude_flip_applied", False)),
                "group": group,
                "variable": variable,
                "unit": unit,
                "biocube_observed_source": "native_target_original.pt" if manifest.get("native_target_original") else "",
                "external_validation_source": "ARPA/manuale",
                "arpa_observed_value": None,
                "arpa_difference": None,
                "arpa_notes": None,
                "within_3pct_biocube": None if error_pct is None else bool(error_pct <= 3.0),
                **summary,
            }
        )
    return pd.DataFrame.from_records(rows)


def update_history_workbook(history_path: Path, latest_area_summary: pd.DataFrame) -> Path:
    """Aggiorna un Excel cumulativo senza perdere i run precedenti."""
    history_path.parent.mkdir(parents=True, exist_ok=True)
    existing_sheets: dict[str, pd.DataFrame] = {}
    if history_path.exists():
        existing_sheets = pd.read_excel(history_path, sheet_name=None)

    previous = existing_sheets.get("monthly_area_summary", pd.DataFrame())
    combined = pd.concat([previous, latest_area_summary], ignore_index=True)
    if not combined.empty:
        combined = combined.drop_duplicates(
            subset=["run_dir", "area_label", "group", "variable"],
            keep="last",
        ).sort_values(["forecast_month", "area_label", "variable"], na_position="last")

    notes = pd.DataFrame.from_records(
        [
            {
                "note": "Questo workbook viene aggiornato a ogni run mensile. Le colonne ARPA sono lasciate vuote per inserire il valore osservato esterno e confrontarlo con BioCube/BioAnalyst.",
            },
            {
                "note": "La colonna observed_* usa il target BioCube del mese forecastato; non e automaticamente un dato ARPA.",
            },
            {
                "note": "Per validazione scientifica confrontare stessa area, stesso mese, stessa unita e stessa definizione temporale.",
            },
        ]
    )

    try:
        with pd.ExcelWriter(history_path) as writer:
            combined.to_excel(writer, sheet_name="monthly_area_summary", index=False)
            latest_area_summary.to_excel(writer, sheet_name="latest_run", index=False)
            notes.to_excel(writer, sheet_name="notes", index=False)
    except PermissionError:
        fallback = history_path.with_name(f"{history_path.stem}_new.xlsx")
        with pd.ExcelWriter(fallback) as writer:
            combined.to_excel(writer, sheet_name="monthly_area_summary", index=False)
            latest_area_summary.to_excel(writer, sheet_name="latest_run", index=False)
            notes.to_excel(writer, sheet_name="notes", index=False)
        return fallback

    return history_path


def export_maps_and_matrix(
    run_dir: Path,
    manifest: dict[str, Any],
    predicted_batch: Any,
    observed_batch: Any | None,
    group: str,
    variable: str,
    area_specs: list[dict[str, Any]],
    matrix_export_format: str = "excel",
) -> tuple[dict[str, str], pd.DataFrame]:
    plots_dir = run_dir / "plots"
    export_dir = run_dir / "exports"
    export_dir.mkdir(parents=True, exist_ok=True)

    predicted_batch = move_tensors_to_cpu(predicted_batch)
    observed_batch = move_tensors_to_cpu(observed_batch) if observed_batch is not None else None

    predicted_map_raw, unit = get_batch_variable_map(predicted_batch, group_name=group, variable_name=variable)
    latitudes, longitudes = get_grid_coordinates(predicted_batch)
    bounds = manifest["bounds"]
    latitude_flip = prediction_latitude_flip_enabled(manifest)
    longitude_flip = prediction_longitude_flip_enabled(manifest)
    predicted_map = align_prediction_map(
        predicted_map_raw,
        latitude_flip=latitude_flip,
        longitude_flip=longitude_flip,
    )

    outputs: dict[str, str] = {}
    prediction_png = plots_dir / f"{variable}_prediction.png"
    plot_map(prediction_png, predicted_map, latitudes, longitudes, f"{group}.{variable} prediction", unit, bounds)
    outputs["prediction_png"] = str(prediction_png)

    observed_map = None
    alignment_diagnostic_frame = None
    if observed_batch is not None:
        observed_map, _ = get_batch_variable_map(observed_batch, group_name=group, variable_name=variable)
        alignment_diagnostic_frame = build_alignment_diagnostic_frame(predicted_map_raw, observed_map)
        observed_png = plots_dir / f"{variable}_observed.png"
        difference_png = plots_dir / f"{variable}_difference.png"
        plot_map(observed_png, observed_map, latitudes, longitudes, f"{group}.{variable} observed", unit, bounds)
        plot_map(difference_png, predicted_map - observed_map, latitudes, longitudes, f"{group}.{variable} prediction - observed", unit, bounds)
        outputs["observed_png"] = str(observed_png)
        outputs["difference_png"] = str(difference_png)

    matrix_outputs, area_summary_frame = export_feature_matrix_bundle(
        run_dir=run_dir,
        export_dir=export_dir,
        workbook_prefix=variable,
        area_summary_prefix=variable,
        manifest=manifest,
        latitudes=latitudes,
        longitudes=longitudes,
        predicted_map=predicted_map,
        observed_map=observed_map,
        predicted_map_raw=predicted_map_raw,
        group=group,
        variable=variable,
        unit=unit,
        area_specs=area_specs,
        matrix_export_format=matrix_export_format,
    )
    outputs.update(matrix_outputs)
    return outputs, area_summary_frame


def export_full_native_output(
    *,
    run_dir: Path,
    manifest: dict[str, Any],
    predicted_batch: Any,
    observed_batch: Any | None,
    selected_group: str,
    selected_variable: str,
    export_group_csvs: bool,
) -> dict[str, str]:
    """Esporta il contenuto del .pt nativo in Excel/CSV senza ricalcolare il modello."""
    from bioanalyst_native_utils import NATIVE_GROUP_FIELDS, flatten_batch_timestamps, list_native_group_counts, list_native_group_variables
    from export_native_output import (
        as_1d_float_array,
        build_grid_frame,
        build_group_wide_frame,
        build_selected_variable_frame,
        build_variable_inventory,
        rows_from_mapping,
        write_excel_with_fallback,
    )

    export_dir = run_dir / "exports"
    export_dir.mkdir(parents=True, exist_ok=True)
    latitudes = as_1d_float_array(predicted_batch.batch_metadata.latitudes)
    longitudes = as_1d_float_array(predicted_batch.batch_metadata.longitudes)
    bounds = manifest.get("bounds")
    group_counts = list_native_group_counts(predicted_batch)
    group_variables = list_native_group_variables(predicted_batch)
    latitude_flip = prediction_latitude_flip_enabled(manifest)
    longitude_flip = prediction_longitude_flip_enabled(manifest)

    run_info = {
        "native_prediction_original": manifest.get("native_prediction_original"),
        "native_target_original": manifest.get("native_target_original"),
        "label": manifest.get("label"),
        "mode": manifest.get("mode"),
        "checkpoint_kind": manifest.get("checkpoint_kind"),
        "input_mode": manifest.get("input_mode"),
        "device": manifest.get("device"),
        "timestamps": flatten_batch_timestamps(predicted_batch),
        "grid_lat_count": int(latitudes.size),
        "grid_lon_count": int(longitudes.size),
        "grid_cell_count": int(latitudes.size * longitudes.size),
        "latitude_min": float(latitudes.min()),
        "latitude_max": float(latitudes.max()),
        "longitude_min": float(longitudes.min()),
        "longitude_max": float(longitudes.max()),
        "selected_group_for_preview": selected_group,
        "selected_variable_for_preview": selected_variable,
        "prediction_latitude_flip_applied": bool(latitude_flip),
        "prediction_longitude_flip_applied": bool(longitude_flip),
        "native_pt_preserved": True,
        "values_note": "Readable exports keep native units and apply declared spatial alignment; native .pt artifacts are preserved unchanged.",
    }
    group_rows = [
        {
            "group": group_name,
            "variable_count": group_counts.get(group_name, 0),
            "variables": ", ".join(group_variables.get(group_name, [])),
        }
        for group_name in NATIVE_GROUP_FIELDS
    ]
    sheets = {
        "run_info": pd.DataFrame(rows_from_mapping(run_info)),
        "groups": pd.DataFrame.from_records(group_rows),
        "variables": build_variable_inventory(predicted_batch, time_index=-1, level_index=0),
        "selected_variable": build_selected_variable_frame(
            predicted_batch,
            group_name=selected_group,
            variable_name=selected_variable,
            latitudes=latitudes,
            longitudes=longitudes,
            bounds=bounds,
            time_index=-1,
            level_index=0,
            observed_batch=observed_batch,
            align_prediction_latitude=latitude_flip,
            align_prediction_longitude=longitude_flip,
        ),
        "coordinates": build_grid_frame(latitudes, longitudes, bounds),
        "manifest": pd.DataFrame(rows_from_mapping(manifest)),
    }

    full_xlsx = write_excel_with_fallback(export_dir / "bioanalyst_native_full_output.xlsx", sheets)
    outputs = {"native_full_xlsx": str(full_xlsx)}

    if export_group_csvs:
        csv_dir = export_dir / "bioanalyst_native_full_output_group_csv"
        csv_dir.mkdir(parents=True, exist_ok=True)
        xlsx_dir = export_dir / "bioanalyst_native_full_output_group_xlsx"
        xlsx_dir.mkdir(parents=True, exist_ok=True)
        for group_name in NATIVE_GROUP_FIELDS:
            group_frame = build_group_wide_frame(
                predicted_batch,
                group_name=group_name,
                latitudes=latitudes,
                longitudes=longitudes,
                bounds=bounds,
                time_index=-1,
                level_index=0,
                align_prediction_latitude=latitude_flip,
                align_prediction_longitude=longitude_flip,
            )
            write_excel_friendly_csv(csv_dir / f"{group_name}_variables_grid.csv", group_frame)
            with pd.ExcelWriter(xlsx_dir / f"{group_name}_variables_grid.xlsx") as writer:
                group_frame.to_excel(writer, sheet_name="full_grid", index=False)
        outputs["native_group_csv_dir"] = str(csv_dir)
        outputs["native_group_xlsx_dir"] = str(xlsx_dir)

    return outputs


def main() -> None:
    args = build_parser().parse_args()
    if args.attention_chunk_size:
        os.environ["BFM_ATTENTION_CHUNK_SIZE"] = str(args.attention_chunk_size)
        install_chunked_cuda_attention(args.attention_chunk_size)

    env = prepare_native_forecast_environment()
    context = build_native_run_context(
        args=args,
        project_output_dir=env["project_output_dir"],
        model_dir=env["model_dir"],
        source_paths=env["source_paths"],
        run_suffix="native_one_step",
    )
    if not context.months_info["compare_available"] and not args.no_compare_observed:
        print(
            "[warn] Il forecast e costruibile, ma il target osservato del mese previsto non e disponibile "
            "con le sorgenti locali correnti. Il run procedera senza confronto observed.",
            flush=True,
        )
    compare_month = (
        context.months_info["forecast_month"]
        if context.months_info["compare_available"] and not args.no_compare_observed
        else None
    )

    print(f"[1/4] Preparo batch native per {context.label}", flush=True)
    saved_windows = prepare_native_saved_windows(
        context=context,
        source_paths=env["source_paths"],
        compare_month=compare_month,
        use_atmospheric_data=not args.fast_smoke_test,
    )

    print(f"[2/4] Carico modello {args.checkpoint} su {context.device}", flush=True)
    runtime = build_native_runtime(
        batch_dir=context.batch_dir,
        checkpoint_path=context.checkpoint_path,
        device=context.device,
    )

    print("[3/4] Eseguo forecast native", flush=True)
    run_kwargs = {
        "context": context,
        "runtime": runtime,
        "saved_windows": saved_windows,
    }
    if "use_amp_bf16" in inspect.signature(run_native_one_step).parameters:
        run_kwargs["use_amp_bf16"] = args.amp_bf16
    elif args.amp_bf16:
        print("[warn] --amp-bf16 ignorato: la versione locale di bioanalyst_native_utils.py non lo supporta.", flush=True)
    result = run_native_one_step(**run_kwargs)
    artifact_paths = save_native_one_step_artifacts(
        context=context,
        runtime=runtime,
        result=result,
    )
    (context.run_dir / "forecast_native_config.yaml").write_text(OmegaConf.to_yaml(runtime.cfg), encoding="utf-8")

    manifest = json.loads(artifact_paths["manifest"].read_text(encoding="utf-8"))
    manifest["lab_export"] = {
        "group": args.group,
        "variable": args.variable,
        "amp_bf16": bool(args.amp_bf16),
        "attention_chunk_size": args.attention_chunk_size,
        "fast_smoke_test": bool(args.fast_smoke_test),
        "input_mode": args.input_mode,
    }
    manifest["spatial_alignment"] = build_spatial_alignment_metadata()
    artifact_paths["manifest"].write_text(json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8")

    print("[4/4] Esporto PNG e matrice Excel", flush=True)
    area_specs = build_area_specs(args, manifest)
    previsioni_run_output_dir = previsioni_month_dir(context.run_dir, manifest)
    export_paths, area_summary_frame = export_maps_and_matrix(
        context.run_dir,
        manifest,
        result.predicted_batch_original,
        result.observed_batch_original,
        args.group,
        args.variable,
        area_specs,
        matrix_export_format=args.matrix_export_format,
    )
    export_paths["reliable_feature_plots"] = export_reliable_feature_plots(
        context.run_dir,
        manifest,
        result.predicted_batch_original,
        result.observed_batch_original,
    )
    print("[extra] Esporto workbook Excel aggiuntivi per feature principali", flush=True)
    export_paths["reliable_feature_workbooks"] = export_reliable_feature_workbooks(
        context.run_dir,
        manifest,
        result.predicted_batch_original,
        result.observed_batch_original,
        area_specs,
        export_root=previsioni_run_output_dir,
        flat_output=True,
        matrix_export_format=args.matrix_export_format,
    )
    print("[extra] Esporto workbook Excel specie", flush=True)
    export_paths["species_matrix_workbook"] = export_species_matrix_workbook(
        context.run_dir,
        manifest,
        result.predicted_batch_original,
        result.observed_batch_original,
        area_specs,
        export_root=previsioni_run_output_dir,
        flat_output=True,
        matrix_export_format=args.matrix_export_format,
    )
    if not args.no_history:
        default_history_name = (
            "native_lab_monthly_results_large.xlsx"
            if args.checkpoint == "large"
            else "native_lab_monthly_results.xlsx"
        )
        history_xlsx = args.history_xlsx or (context.run_dir.parent / default_history_name)
        export_paths["history_xlsx"] = str(update_history_workbook(history_xlsx, area_summary_frame))

    if args.export_native_full:
        print("[extra] Esporto output nativo completo BioAnalyst in Excel/CSV", flush=True)
        export_paths.update(
            export_full_native_output(
                run_dir=context.run_dir,
                manifest=manifest,
                predicted_batch=result.predicted_batch_original,
                observed_batch=result.observed_batch_original,
                selected_group=args.group,
                selected_variable=args.variable,
                export_group_csvs=args.export_native_group_csvs,
            )
        )

    if not args.no_biomap_final_workbook:
        print("[extra] Aggiorno workbook finale BIOMAP", flush=True)
        biomap_workbook_path = (
            args.biomap_final_workbook
            if args.biomap_final_workbook is not None
            else previsioni_root_for_run(context.run_dir) / DEFAULT_WORKBOOK_NAME
        )
        export_paths["biomap_final_workbook"] = update_biomap_final_workbook(
            current_run_dir=context.run_dir,
            prediction_batch=result.predicted_batch_original,
            observed_batch=result.observed_batch_original,
            manifest=manifest,
            workbook_path=biomap_workbook_path,
        )
        workbook_snapshot_path = previsioni_run_output_dir / DEFAULT_WORKBOOK_NAME
        workbook_snapshot_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(export_paths["biomap_final_workbook"]["workbook"], workbook_snapshot_path)
        export_paths["biomap_final_workbook"]["workbook_snapshot"] = str(workbook_snapshot_path)

    print(
        json.dumps(
            {
                "run_dir": str(context.run_dir),
                "manifest": str(artifact_paths["manifest"]),
                "native_prediction_original": str(artifact_paths["prediction"]),
                "native_target_original": str(artifact_paths["observed"]) if artifact_paths["observed"] else None,
                "exports": export_paths,
            },
            indent=2,
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
